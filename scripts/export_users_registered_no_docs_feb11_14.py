#!/usr/bin/env python3
"""
Экспорт в Excel: клиенты, зарегистрировавшиеся 11–14 февраля, не загрузившие документы.
Запуск из корня проекта: python scripts/export_users_registered_no_docs_feb11_14.py
Требуется: pip install openpyxl
"""
import os
import sys
from pathlib import Path

# Корень проекта = родитель папки scripts
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
os.chdir(ROOT)

from dotenv import load_dotenv
load_dotenv()

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("Установите openpyxl: pip install openpyxl")
    sys.exit(1)

from sqlalchemy import create_engine, text
from app.core.config import DATABASE_URL


def main():
    # Заменить на psycopg2 URL для синхронного подключения (create_engine по умолчанию sync)
    engine = create_engine(DATABASE_URL.replace("postgresql+psycopg2", "postgresql+psycopg2"))
    # Даты 11–14 февраля по Алматы (UTC+5). Если в БД created_at хранится в UTC — раскомментируйте вариант с AT TIME ZONE.
    sql = """
    SELECT
        u.phone_number,
        u.email,
        TRIM(COALESCE(u.first_name, '') || ' ' || COALESCE(u.last_name, '') || ' ' || COALESCE(u.middle_name, '')) AS full_name,
        u.first_name,
        u.last_name,
        u.middle_name,
        u.created_at
    FROM users u
    WHERE (
        (u.created_at::date IN ('2026-02-11', '2026-02-12', '2026-02-13', '2026-02-14'))
    )
      AND u.upload_document_at IS NULL
      AND u.is_deleted = false
    ORDER BY u.created_at, u.phone_number
    """
    with engine.connect() as conn:
        result = conn.execute(text(sql))
        rows = result.fetchall()

    # Заголовки для Excel (русские)
    headers = ["Номер", "Email", "ФИО", "Имя", "Фамилия", "Отчество", "Дата регистрации"]
    out_path = ROOT / "scripts" / "users_registered_no_docs_11_14_feb.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Клиенты без документов"

    # Заголовок
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    # Данные (время в Excel — по Алматы, с подписью)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            if hasattr(val, "isoformat"):
                if val:
                    # Формат даты/времени по Алматы (в БД уже хранится как локальное Алматы, UTC+5)
                    val = val.strftime("%Y-%m-%d %H:%M:%S") + " (Алматы)"
                else:
                    val = ""
            ws.cell(row=row_idx, column=col_idx, value=val)
    # Ширина колонок
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18

    wb.save(out_path)
    print(f"Сохранено: {out_path}")
    print(f"Строк (без заголовка): {len(rows)}")


if __name__ == "__main__":
    main()
